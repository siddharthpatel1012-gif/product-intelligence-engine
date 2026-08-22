"""
Batch File I/O
---------------
Reads an uploaded CSV or XLSX of raw input rows into plain dicts, and
writes processed rows back out as a single XLSX with the exact fixed
header row required by the hackathon's Expected Output sheet.

No pandas dependency — openpyxl (already lightweight, one new
requirement) for XLSX, the stdlib csv module for CSV. Keeps the
deploy footprint small on a free-tier host.
"""
import csv
import io
from openpyxl import Workbook, load_workbook


def read_input_file(content: bytes, filename: str) -> list[dict]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _read_csv(content)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_xlsx(content)
    raise ValueError(f"Unsupported file type: {filename}. Upload a .csv or .xlsx file.")


def _read_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")  # handles a leading BOM from Excel exports
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    all_rows_raw = [r for r in rows_iter if not (r is None or all(v is None for v in r))]

    # Defensive fallback for a common Excel mistake: pasting CSV text
    # into a spreadsheet without splitting it into columns leaves the
    # whole comma-separated line sitting in a single cell (column A)
    # for every row. Detect that shape — one header cell containing
    # commas, one value cell per data row also containing the same
    # number of commas — and re-split on comma rather than silently
    # returning unusable single-column rows.
    if len(headers) == 1 and "," in headers[0]:
        split_headers = [h.strip() for h in headers[0].split(",")]
        looks_like_csv_in_one_cell = all(
            len(row) == 1 and isinstance(row[0], str) and row[0].count(",") == headers[0].count(",")
            for row in all_rows_raw
        ) if all_rows_raw else True
        if looks_like_csv_in_one_cell:
            headers = split_headers
            all_rows_raw = [tuple(v.split(",")) for (v,) in all_rows_raw]

    rows = []
    for raw in all_rows_raw:
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        rows.append(row)
    return rows


def write_xlsx_bytes(rows: list[dict], headers: list[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Output"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_csv_bytes(rows: list[dict], headers: list[str]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({h: row.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8-sig")